import streamlit as st
import zipfile
import pandas as pd
import openpyxl
from openpyxl import Workbook
from io import BytesIO
from datetime import datetime
import re

st.set_page_config(page_title="Bank File Merger v2.0 | 28/02 08:00", page_icon="🏦", layout="wide")

# ── BANK PROFILES ──────────────────────────────────────────────
def detect_bank(rows):
    flat = ' '.join([str(c) for r in rows[:15] for c in r if c])
    if 'BẢNG SAO KÊ GIAO DỊCH' in flat:
        return 'ACB'
    if 'SAO KÊ TÀI KHOẢN' in flat or 'STATEMENT OF ACCOUNT' in flat:
        return 'VCB'
    if 'so but toan' in flat.lower() and 'ngay giao dich' in flat.lower():
        return 'TCB'
    if 'VIETINBANK' in flat.upper() or 'efast' in flat.lower() or 'LỊCH SỬ GIAO DỊCH' in flat:
        return 'VTB'
    if 'MB BANK' in flat.upper() or 'MILITARY' in flat.upper():
        return 'MB'
    return None

def get_account_no(rows, bank_id):
    flat_rows = [' '.join([str(c) for c in r if c]) for r in rows[:15]]
    if bank_id == 'ACB':
        for line in flat_rows:
            m = re.search(r'[Ss]ố tài khoản.*?:\s*(\d+)', line)
            if m: return m.group(1)
            m = re.search(r'[Tt]ài khoản số:\s*(\d+)', line)
            if m: return m.group(1)
    elif bank_id == 'VCB':
        for r in rows[:10]:
            for i, cell in enumerate(r):
                if cell and ('tài khoản' in str(cell).lower() or 'account number' in str(cell).lower()):
                    for j in range(i+1, len(r)):
                        v = str(r[j] or '').strip()
                        if re.match(r'^\d{8,}$', v): return v
                    m = re.search(r'\d{8,}', str(cell))
                    if m: return m.group(0)
    elif bank_id == 'TCB':
        if len(rows) > 1 and len(rows[1]) > 1:
            return str(rows[1][1] or '').strip()
    elif bank_id in ('VTB', 'MB'):
        for r in rows[:15]:
            for i, cell in enumerate(r):
                if cell and ('account no' in str(cell).lower() or 'số tài khoản' in str(cell).lower()):
                    if i+1 < len(r):
                        m = re.search(r'\d{8,}', str(r[i+1] or ''))
                        if m: return m.group(0)
    return 'unknown'

def find_header_row(rows, bank_id):
    kws = {
        'ACB': ['ngày hiệu lực', 'số gd'],
        'VCB': ['debit', 'credit'],
        'TCB': ['so but toan', 'no/debit'],
        'VTB': ['accounting date', 'debit'],
        'MB':  ['ngày giao dịch', 'số tiền'],
    }
    keywords = kws.get(bank_id, [])
    for i, row in enumerate(rows):
        # Normalize: replace newlines + tabs → space trước khi so sánh
        flat = ' '.join([str(c or '').replace('\n',' ').replace('\t',' ').lower() for c in row])
        if all(kw in flat for kw in keywords):
            return i
    return -1

def parse_amount(val):
    """Normalize số: xóa dấu . và , phân cách nghìn → số nguyên"""
    if val is None or str(val).strip() == '': return 0
    s = str(val).strip()
    # Xóa chữ VND và ký tự không phải số ở cuối (VD: "21,991,508 VND")
    s = re.sub(r'[A-Za-z\s]+$', '', s).strip()
    if not s: return 0
    # Xóa tất cả dấu chấm và phẩy (VN dùng . hoặc , để phân cách nghìn)
    s = re.sub(r'[,\.]', '', s)
    try:
        return int(float(s))
    except:
        return 0

def parse_date(val):
    """Parse date từ nhiều format"""
    if not val: return None
    s = str(val).strip().split('\n')[0]  # VCB merged cell
    # Nếu là datetime object (openpyxl trả về datetime)
    if hasattr(val, 'year'): return val
    patterns = [
        r'^(\d{1,2})/(\d{1,2})/(\d{4})',           # dd/mm/yyyy
        r'^(\d{1,2})/(\d{1,2})/(\d{4})\s+\d{1,2}:\d{2}', # dd/mm/yyyy HH:MM
        r'^(\d{4})-(\d{2})-(\d{2})',                 # yyyy-mm-dd
        r'^(\d{1,2})-(\d{1,2})-(\d{4})',             # dd-mm-yyyy
        r'^(\d{1,2})-(\d{1,2})-(\d{4})\s+\d{1,2}:\d{2}', # dd-mm-yyyy HH:MM:SS
    ]
    for p in patterns:
        m = re.match(p, s)
        if m:
            g = m.groups()
            try:
                if len(g[0]) == 4:  # yyyy-mm-dd
                    return datetime(int(g[0]), int(g[1]), int(g[2]))
                else:
                    return datetime(int(g[2]), int(g[1]), int(g[0]))
            except:
                continue
    return None

def get_dedup_key(row, headers, bank_id, account_no):
    """Tạo key để dedup"""
    h = [str(h or '').lower() for h in headers]

    # Tìm Số GD / reference
    ref = ''
    for kw in ['số gd', 'so but toan', 'số giao dịch', 'reference', 'số tham chiếu']:
        for i, hh in enumerate(h):
            if kw in hh and i < len(row):
                ref = str(row[i] or '').strip()
                break
        if ref: break

    # Tìm ngày
    date_str = ''
    for kw in ['ngày giao dịch', 'ngay giao dich', 'ngày hạch toán', 'transaction date']:
        for i, hh in enumerate(h):
            if kw in hh and i < len(row):
                date_str = str(row[i] or '').strip()
                break
        if date_str: break
    if not date_str and len(row) > 0:
        date_str = str(row[0] or '').strip()

    # Tìm số tiền
    amounts = []
    for kw in ['tiền', 'debit', 'credit', 'nợ', 'có', 'rút', 'gửi', 'no/', 'co/']:
        for i, hh in enumerate(h):
            if kw in hh and i < len(row):
                v = parse_amount(row[i])
                if v > 0: amounts.append(str(v))

    if ref:
        return f"{bank_id}_{account_no}_{ref}"
    else:
        return f"{bank_id}_{account_no}_{date_str}_{'|'.join(amounts)}"

def normalize_row(row, headers):
    """Normalize số trong row"""
    result = []
    h = [str(h or '').lower() for h in headers]
    for i, cell in enumerate(row):
        col_name = h[i] if i < len(h) else ''
        is_amount = any(kw in col_name for kw in [
            'tiền', 'nợ', 'có', 'debit', 'credit', 'dư', 'balance',
            'rút', 'gửi', 'no/', 'co/', 'amount'
        ])
        if is_amount:
            result.append(parse_amount(cell))
        else:
            result.append(str(cell) if cell is not None else '')
    return result

def read_file(uploaded_file):
    """Đọc file xlsx/xls/csv → list of rows"""
    name = uploaded_file.name.lower()
    if name.endswith('.csv'):
        # Auto-detect separator: thử ; trước rồi ,
        raw = uploaded_file.read()
        # Detect encoding (handle BOM)
        for enc in ['utf-8-sig', 'utf-8', 'latin-1', 'cp1252']:
            try:
                text = raw.decode(enc)
                break
            except:
                continue
        # Parse thủ công từng dòng để tránh lỗi pandas với file có số cột không đều
        lines = text.replace('\r\n', '\n').replace('\r', '\n').split('\n')
        rows = []
        for line in lines:
            if not line.strip():
                rows.append([])
                continue
            # Detect separator từ dòng có nhiều field nhất
            sep = ';' if line.count(';') > line.count(',') else ','
            # Parse thủ công handle quoted fields
            cols = []
            cur = ''
            in_q = False
            for ch in line:
                if ch == '"':
                    in_q = not in_q
                elif ch == sep and not in_q:
                    cols.append(cur.strip())
                    cur = ''
                else:
                    cur += ch
            cols.append(cur.strip())
            rows.append(cols)
        return rows
    elif name.endswith('.xls'):
        # Format cũ Excel 97-2003 → dùng xlrd
        import xlrd
        wb = xlrd.open_workbook(file_contents=uploaded_file.read())
        ws = wb.sheet_by_index(0)
        rows = []
        for i in range(ws.nrows):
            rows.append([ws.cell_value(i, j) for j in range(ws.ncols)])
        return rows
    else:
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        return rows

def process_files(files_by_group):
    """Merge + dedup files theo nhóm"""
    results = {}
    for key, info in files_by_group.items():
        bank_id = info['bank_id']
        account_no = info['account_no']
        all_rows_data = info['files']  # list of (rows, filename)

        if not all_rows_data:
            continue

        # Lấy header từ file đầu tiên
        first_rows = all_rows_data[0][0]
        h_idx = find_header_row(first_rows, bank_id)
        if h_idx < 0:
            results[key] = {'error': f'Không tìm thấy header row trong file {all_rows_data[0][1]}'}
            continue

        meta_rows = first_rows[:h_idx]
        header_row = first_rows[h_idx]
        headers = header_row

        # Gom tất cả data rows
        seen = set()
        all_data = []
        total_input = 0
        dup_count = 0

        for rows, fname in all_rows_data:
            this_h = find_header_row(rows, bank_id)
            if this_h < 0: continue

            for row in rows[this_h+1:]:
                # Bỏ qua row rỗng
                flat = ''.join([str(c or '') for c in row]).strip()
                if not flat: continue

                # Check có ngày hợp lệ không - tìm trong các col đầu
                d = None
                for _ci in range(min(5, len(row))):
                    d = parse_date(row[_ci])
                    if d: break
                if not d: continue

                total_input += 1

                # Dedup
                dk = get_dedup_key(row, headers, bank_id, account_no)
                if dk in seen:
                    dup_count += 1
                    continue
                seen.add(dk)

                # Normalize
                clean_row = normalize_row(row, headers)
                all_data.append((d, clean_row))

        # Sort theo ngày tăng dần
        all_data.sort(key=lambda x: x[0])

        if not all_data:
            results[key] = {'error': 'Không có data sau khi lọc'}
            continue

        # Date range cho tên file
        min_date = all_data[0][0]
        max_date = all_data[-1][0]
        fname = f"{bank_id}_{account_no}_{min_date.strftime('%d%m%Y')}to{max_date.strftime('%d%m%Y')}.xlsx"

        # Build output workbook
        wb_out = Workbook()
        ws_out = wb_out.active
        for r in meta_rows:
            ws_out.append([c if c is not None else '' for c in r])
        ws_out.append([c if c is not None else '' for c in header_row])
        for _, row in all_data:
            ws_out.append(row)

        buf = BytesIO()
        wb_out.save(buf)
        buf.seek(0)

        results[key] = {
            'filename': fname,
            'data': buf,
            'tx_count': len(all_data),
            'dup_removed': dup_count,
            'total_input': total_input,
            'date_from': min_date.strftime('%d/%m/%Y'),
            'date_to': max_date.strftime('%d/%m/%Y'),
        }

    return results

# ── UI ─────────────────────────────────────────────────────────
st.title("🏦 Bank File Merger v2.0 | 28/02 08:00")
st.caption("Upload file sao kê ngân hàng → Tự nhận dạng → Merge + Dedup → Xuất file sạch")

uploaded = st.file_uploader(
    "📂 Upload file sao kê (xlsx, xls, csv) — chọn nhiều file cùng lúc",
    type=['xlsx', 'xls', 'csv'],
    accept_multiple_files=True
)

if uploaded:
    st.divider()

    # Phân nhóm file theo ngân hàng + số TK
    groups = {}
    errors = []

    with st.spinner("🔍 Đang nhận dạng file..."):
        for f in uploaded:
            try:
                rows = read_file(f)
                bank_id = detect_bank(rows)
                if not bank_id:
                    errors.append(f"❓ **{f.name}** — Không nhận dạng được ngân hàng")
                    continue
                account_no = get_account_no(rows, bank_id)
                key = f"{bank_id}_{account_no}"
                if key not in groups:
                    groups[key] = {'bank_id': bank_id, 'account_no': account_no, 'files': []}
                groups[key]['files'].append((rows, f.name))
            except Exception as e:
                errors.append(f"❌ **{f.name}** — Lỗi: {str(e)}")

    # Hiển thị lỗi
    if errors:
        with st.expander("⚠️ File không nhận dạng được", expanded=True):
            for e in errors:
                st.markdown(e)

    if not groups:
        st.warning("Không có file nào được nhận dạng.")
        st.stop()

    # Hiển thị nhóm
    st.subheader(f"📊 Tìm thấy {len(groups)} nhóm từ {len(uploaded)} file")

    cols = st.columns(min(len(groups), 4))
    for i, (key, info) in enumerate(groups.items()):
        with cols[i % len(cols)]:
            st.metric(
                label=f"{info['bank_id']} · {info['account_no']}",
                value=f"{len(info['files'])} file",
            )

    st.divider()

    # Nút merge
    if st.button("⚡ Merge & Dedup tất cả", type="primary", use_container_width=True):
        with st.spinner("⏳ Đang xử lý..."):
            results = process_files(groups)

        # Lưu vào session_state để Phase 2 dùng được
        st.session_state.merge_results = results

        st.success(f"✅ Hoàn tất! {len(results)} file đã được tạo")
        st.divider()

        # Nút Download All - zip tất cả file
        ok_results = {k:v for k,v in results.items() if 'error' not in v}
        if len(ok_results) > 1:
            zip_buf = BytesIO()
            with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                for k, r in ok_results.items():
                    r['data'].seek(0)
                    zf.writestr(r['filename'], r['data'].read())
            zip_buf.seek(0)
            st.download_button(
                label=f"⬇️ Tải tất cả ({len(ok_results)} file) — ZIP",
                data=zip_buf,
                file_name="bank_merged_all.zip",
                mime="application/zip",
                type="primary",
                use_container_width=True,
                key="dl_all"
            )
            st.divider()

        for key, res in results.items():
            if 'error' in res:
                st.error(f"❌ **{key}**: {res['error']}")
                continue

            col1, col2 = st.columns([3, 1])
            with col1:
                st.markdown(f"**📄 {res['filename']}**")
                st.caption(
                    f"✅ {res['tx_count']} giao dịch | "
                    f"🗑️ Bỏ {res['dup_removed']} trùng | "
                    f"📅 {res['date_from']} → {res['date_to']}"
                )
            with col2:
                st.download_button(
                    label="⬇️ Tải về",
                    data=res['data'],
                    file_name=res['filename'],
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    key=f"dl_{key}"
                )


# ═══════════════════════════════════════════════════════════
# PHASE 2 — KẾT NỐI GOOGLE SHEETS + DUYỆT LỆNH
# ═══════════════════════════════════════════════════════════
import json
from google.oauth2.service_account import Credentials
import gspread

SPREADSHEET_ID = '1ykPA0eFSJKjcK1ofRA4ZFD5YtqoWHgfzFnCoWXysSUU'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets', 'https://www.googleapis.com/auth/drive']

# Các sheet KHÔNG phải dự án
NON_PROJECT_SHEETS = [
    'Account', 'Banking', 'Staging', 'Config',
    'Bán hàng EZ', 'Cac sim EZ', 'Mua bán EZ',
    'EZ 5335 chi ra', 'EZ 5335 nhan lai',
    'Bai tap ve nha', 'Doi soat VNSKY', 'Dong tien bank',
    'Doi soat Phat', 'Doi soat A Minh', 'Chia co tuc T10.2021',
    'Danh sach ma diem ban Khoa', 'Final T6_CT2',
]

# Raw sheet → Account sheet row mapping
RAW_TO_ACCOUNT = {
    '3651368':    {'name': 'Raw_ACB_Khoa',     'bank': 'ACB'},
    '18091368':   {'name': 'Raw_ACB_NamKhang', 'bank': 'ACB'},
    '1112136888': {'name': 'Raw_ACB_VTSG',     'bank': 'ACB'},
    '0721000656789': {'name': 'Raw_VCB_Khoa',  'bank': 'VCB'},
    'Vietin VTSG':   {'name': 'Raw_Vietin_VTSG', 'bank': 'VTB'},
    'Vietin Naka':   {'name': 'Raw_Vietin_Naka', 'bank': 'VTB'},
    'Tech NAKA':     {'name': 'Raw_Tech_Naka',   'bank': 'TCB'},
}

def connect_gsheet(creds_json):
    """Kết nối Google Sheets từ credentials JSON"""
    try:
        creds = Credentials.from_service_account_info(creds_json, scopes=SCOPES)
        client = gspread.authorize(creds)
        spreadsheet = client.open_by_key(SPREADSHEET_ID)
        return spreadsheet, None
    except Exception as e:
        return None, str(e)

def get_sheet_names(spreadsheet):
    """Lấy tất cả sheet names"""
    return [ws.title for ws in spreadsheet.worksheets()]

def get_project_sheets(spreadsheet):
    """Lấy các sheet dự án (loại trừ raw + non-project)"""
    all_sheets = get_sheet_names(spreadsheet)
    raw_names = [v['name'] for v in RAW_TO_ACCOUNT.values()]
    raw_names += list(RAW_TO_ACCOUNT.keys())
    project = [s for s in all_sheets
               if s not in NON_PROJECT_SHEETS
               and s not in raw_names
               and not s.lower().startswith('raw_')]
    return project

def get_sheet_history(spreadsheet, sheet_name, max_rows=200):
    """Lấy lịch sử data của 1 sheet để học pattern"""
    try:
        ws = spreadsheet.worksheet(sheet_name)
        data = ws.get_all_values()
        return data[-max_rows:] if len(data) > max_rows else data
    except:
        return []

def suggest_project_sheet(description, project_sheets, spreadsheet):
    """
    Đề xuất sheet dự án dựa vào nội dung giao dịch
    Logic: tìm sheet nào có nhiều keyword giống description nhất
    """
    if not description:
        return project_sheets[0] if project_sheets else None

    desc_words = set(description.upper().split())
    scores = {}

    for sheet_name in project_sheets:
        history = get_sheet_history(spreadsheet, sheet_name, max_rows=100)
        if not history:
            scores[sheet_name] = 0
            continue

        score = 0
        # Lấy tất cả text từ lịch sử sheet
        all_text = ' '.join([' '.join([str(c) for c in row]) for row in history]).upper()

        # Tìm keyword overlap
        for word in desc_words:
            if len(word) >= 4 and word in all_text:
                score += 1

        # Bonus: tên sheet xuất hiện trong description
        if sheet_name.upper() in description.upper():
            score += 10

        scores[sheet_name] = score

    # Sort theo score cao nhất
    sorted_sheets = sorted(scores.items(), key=lambda x: x[1], reverse=True)
    return sorted_sheets[0][0] if sorted_sheets else project_sheets[0]

def get_account_cell(spreadsheet, raw_sheet_name):
    """Tìm cell số dư trong sheet Account cho raw sheet tương ứng"""
    try:
        ws = spreadsheet.worksheet('Account')
        data = ws.get_all_values()
        for i, row in enumerate(data):
            for j, cell in enumerate(row):
                if str(cell).strip() == raw_sheet_name:
                    # Tìm cell số tiếp theo cùng dòng có giá trị số
                    for k in range(j+1, len(row)):
                        v = str(row[k]).replace(',','').replace('.','').strip()
                        if v.isdigit():
                            # Trả về địa chỉ cell (row i+1, col k+1)
                            col_letter = chr(65 + k)
                            return f"{col_letter}{i+1}", float(row[k].replace(',',''))
    except Exception as e:
        pass
    return None, None

def append_to_raw_sheet(spreadsheet, raw_sheet_name, row_data):
    """Append 1 dòng vào raw sheet"""
    ws = spreadsheet.worksheet(raw_sheet_name)
    ws.append_row(row_data, value_input_option='USER_ENTERED')

def append_to_project_sheet(spreadsheet, project_sheet_name, date_str, desc, amount):
    """Append 1 dòng vào sheet dự án (nghịch dấu với tài khoản)"""
    ws = spreadsheet.worksheet(project_sheet_name)
    ws.append_row([date_str, desc, -amount], value_input_option='USER_ENTERED')

def update_account_balance(spreadsheet, cell_addr, delta):
    """Cộng/trừ số dư tài khoản trong sheet Account"""
    ws = spreadsheet.worksheet('Account')
    current = ws.acell(cell_addr).value
    current_val = float(str(current).replace(',','').replace('.','')) if current else 0
    new_val = current_val + delta
    ws.update_acell(cell_addr, new_val)

def build_raw_row(tx, raw_sheet_name, spreadsheet):
    """Build row data để append vào raw sheet, auto-map columns"""
    try:
        ws = spreadsheet.worksheet(raw_sheet_name)
        header = ws.row_values(1)
        if not header:
            # Default: date, desc, debit, credit, balance, ref
            return [tx.get('date',''), tx.get('desc',''),
                    tx.get('debit',0), tx.get('credit',0),
                    tx.get('balance',''), tx.get('ref','')]

        row = []
        for col in header:
            col_l = col.lower().strip()
            if any(k in col_l for k in ['ngày','date','ngay']):
                row.append(tx.get('date',''))
            elif any(k in col_l for k in ['nội dung','diễn giải','mô tả','desc','noi dung']):
                row.append(tx.get('desc',''))
            elif any(k in col_l for k in ['rút','nợ','debit','ghi nợ','chi']):
                row.append(tx.get('debit',0))
            elif any(k in col_l for k in ['gửi','có','credit','ghi có','thu']):
                row.append(tx.get('credit',0))
            elif any(k in col_l for k in ['số dư','balance','so du']):
                row.append(tx.get('balance',''))
            elif any(k in col_l for k in ['số gd','ref','so gd','but toan']):
                row.append(tx.get('ref',''))
            elif any(k in col_l for k in ['tên tk','tên tài khoản','counter name']):
                row.append(tx.get('counter_name',''))
            elif any(k in col_l for k in ['tk đối','tài khoản đối','counter acc']):
                row.append(tx.get('counter_acct',''))
            else:
                row.append('')
        return row
    except:
        return [tx.get('date',''), tx.get('desc',''),
                tx.get('debit',0), tx.get('credit',0)]

# ── PHASE 2 UI ───────────────────────────────────────────────


BIG_ISSUE_OPTION = "⚡ Big Issue (D86 - Account)"
BIG_ISSUE_CELL = "D86"
BIG_ISSUE_SHEET = "Account"

def update_big_issue(spreadsheet, delta):
    """Cộng/trừ trực tiếp vào cell D86 trong sheet Account"""
    ws = spreadsheet.worksheet(BIG_ISSUE_SHEET)
    current = ws.acell(BIG_ISSUE_CELL).value
    current_val = 0
    if current:
        s = str(current).replace(',','').replace('.','').strip()
        try:
            current_val = float(s)
        except:
            current_val = 0
    new_val = current_val + delta
    ws.update_acell(BIG_ISSUE_CELL, new_val)

def get_last_ref_from_raw(spreadsheet, raw_sheet_name):
    """B2: Lấy ref cuối cùng từ Raw sheet"""
    try:
        ws = spreadsheet.worksheet(raw_sheet_name)
        all_data = ws.get_all_values()
        if not all_data or len(all_data) < 2:
            return None
        
        header = [str(c or '').lower().strip() for c in all_data[0]]
        ref_col_idx = -1
        for i, h in enumerate(header):
            if any(k in h for k in ['số gd', 'so gd', 'ref', 'but toan', 'transaction number', 
                                      'số giao dịch', 'số tham chiếu', 'reference']):
                ref_col_idx = i
                break
        
        if ref_col_idx < 0:
            return None
        
        for row in reversed(all_data[1:]):
            if ref_col_idx < len(row):
                val = str(row[ref_col_idx] or '').strip()
                if val:
                    return val
        return None
    except:
        return None

def get_account_balance_for_raw(spreadsheet, raw_sheet_name):
    """B3: Lấy số dư từ sheet Account cho raw sheet tương ứng"""
    try:
        ws = spreadsheet.worksheet('Account')
        data = ws.get_all_values()
        for i, row in enumerate(data):
            for j, cell in enumerate(row):
                if str(cell).strip() == raw_sheet_name:
                    for k in range(j+1, len(row)):
                        v = str(row[k]).replace(',','').strip()
                        v_clean = v.replace('.','')
                        if v_clean.isdigit() or (v_clean.startswith('-') and v_clean[1:].isdigit()):
                            return int(v_clean)
        return None
    except:
        return None

# ── PHASE 2 UI ───────────────────────────────────────────────
def render_phase2():
    st.divider()
    st.header("📋 Phase 2 — Duyệt lệnh & Hạch toán")

    # Load credentials
    creds_json = None
    try:
        creds_json = dict(st.secrets["gcp_service_account"])
    except:
        creds_file = st.file_uploader(
            "🔑 Upload file credentials JSON (Google Service Account)",
            type=['json'], key='creds_uploader'
        )
        if not creds_file:
            st.info("Upload file credentials JSON để kết nối Google Sheets")
            return
        creds_json = json.load(creds_file)

    # Kết nối (cache)
    if 'gsheet_connected' not in st.session_state:
        with st.spinner("🔌 Đang kết nối Google Sheets..."):
            spreadsheet, err = connect_gsheet(creds_json)
        if err:
            st.error(f"❌ Kết nối thất bại: {err}")
            return
        st.session_state.gsheet_connected = True
        st.session_state.spreadsheet = spreadsheet
        st.session_state.project_sheets = get_project_sheets(spreadsheet)
    
    spreadsheet = st.session_state.spreadsheet
    project_sheets = st.session_state.project_sheets
    
    st.success(f"✅ Đã kết nối: **{spreadsheet.title}**")

    if 'merge_results' not in st.session_state or not st.session_state.merge_results:
        st.warning("⚠️ Chưa có file nào được merge. Vui lòng chạy Phase 1 trước!")
        return

    ok_results = {k:v for k,v in st.session_state.merge_results.items() if 'error' not in v}
    if not ok_results:
        st.warning("Không có file hợp lệ từ Phase 1")
        return

    file_options = list(ok_results.keys())
    selected_key = st.selectbox(
        "📂 Chọn file để duyệt",
        file_options,
        format_func=lambda k: ok_results[k]['filename']
    )

    res = ok_results[selected_key]
    res['data'].seek(0)

    wb = openpyxl.load_workbook(res['data'], data_only=True)
    ws_merged = wb.active
    rows = [list(r) for r in ws_merged.iter_rows(values_only=True)]

    bank_id = selected_key.split('_')[0]
    h_idx = find_header_row(rows, bank_id)
    if h_idx < 0:
        st.error("Không tìm thấy header row trong file merged")
        return

    headers = [str(c or '').replace('\n',' ').strip() for c in rows[h_idx]]

    # Build TOÀN BỘ transactions
    all_transactions = []
    for row in rows[h_idx+1:]:
        flat = ''.join([str(c or '') for c in row]).strip()
        if not flat: continue

        d = None
        date_str = ''
        for ci in range(min(5, len(row))):
            d = parse_date(row[ci])
            if d:
                date_str = str(row[ci]).split('\n')[0].strip()
                break
        if not d: continue

        tx = {'date': date_str, 'desc': '', 'debit': 0, 'credit': 0,
              'balance': 0, 'ref': '', 'counter_name': '', 'counter_acct': ''}

        for i, h in enumerate(headers):
            if i >= len(row): continue
            val = row[i]
            h_l = h.lower()
            if any(k in h_l for k in ['nội dung','diễn giải','mô tả','description','transactions in detail']):
                tx['desc'] = str(val or '').strip()
            elif any(k in h_l for k in ['rút ra','ghi nợ','nợ/ debit','no/debit','debit']):
                tx['debit'] = parse_amount(val)
            elif any(k in h_l for k in ['gửi vào','ghi có','có / credit','co/credit','credit']):
                tx['credit'] = parse_amount(val)
            elif any(k in h_l for k in ['số dư','balance']):
                tx['balance'] = parse_amount(val)
            elif any(k in h_l for k in ['số gd','so but toan','transaction number','số giao dịch','số tham chiếu','reference']):
                tx['ref'] = str(val or '').strip()
            elif any(k in h_l for k in ['tên tk','corresponsive name','tên tài khoản đối']):
                tx['counter_name'] = str(val or '').strip()
            elif any(k in h_l for k in ['tk đối','corresponsive account','số tài khoản đối']):
                tx['counter_acct'] = str(val or '').strip()

        if tx['debit'] == 0 and tx['credit'] == 0: continue
        tx['direction'] = 'THU' if tx['credit'] > 0 else 'CHI'
        tx['amount'] = tx['credit'] if tx['credit'] > 0 else tx['debit']
        all_transactions.append(tx)

    if not all_transactions:
        st.warning("Không có giao dịch nào trong file này")
        return

    # ═══════════════════════════════════════════════
    # B2: TÌM ĐIỂM CẮT
    # ═══════════════════════════════════════════════
    acct_no = selected_key.split('_')[1] if '_' in selected_key else ''
    raw_sheet_candidates = [k for k,v in RAW_TO_ACCOUNT.items() if acct_no in k]
    raw_sheet_key = raw_sheet_candidates[0] if raw_sheet_candidates else acct_no
    raw_sheet_gsheet = RAW_TO_ACCOUNT[raw_sheet_key]['name'] if raw_sheet_key in RAW_TO_ACCOUNT else raw_sheet_key

    st.markdown(f"🏦 **{bank_id}** · `{acct_no}` → Raw sheet: `{raw_sheet_gsheet}`")

    with st.spinner("🔍 B2: Đang tìm giao dịch cuối trong Raw sheet..."):
        last_ref = get_last_ref_from_raw(spreadsheet, raw_sheet_gsheet)

    cutoff_idx = -1
    cutoff_balance = 0
    if last_ref:
        for i, tx in enumerate(all_transactions):
            if tx['ref'] == last_ref:
                cutoff_idx = i
                cutoff_balance = tx['balance']
                break

    if last_ref and cutoff_idx >= 0:
        st.info(f"🔗 Ref cuối trong Raw sheet: `{last_ref}` → vị trí #{cutoff_idx + 1}/{len(all_transactions)}")
        new_transactions = all_transactions[cutoff_idx + 1:]
    elif last_ref and cutoff_idx < 0:
        st.warning(f"⚠️ Ref cuối `{last_ref}` không tìm thấy trong file merged. Hiển thị tất cả.")
        new_transactions = all_transactions
        cutoff_balance = 0
    else:
        st.info("📭 Raw sheet trống — hiển thị tất cả giao dịch")
        new_transactions = all_transactions
        cutoff_balance = 0

    # ═══════════════════════════════════════════════
    # B3: DOUBLE CHECK SỐ DƯ
    # ═══════════════════════════════════════════════
    if cutoff_idx >= 0 and cutoff_balance > 0:
        with st.spinner("🔍 B3: Đang kiểm tra số dư..."):
            account_balance = get_account_balance_for_raw(spreadsheet, raw_sheet_gsheet)

        if account_balance is not None:
            diff = cutoff_balance - account_balance
            
            col_b1, col_b2, col_b3 = st.columns(3)
            with col_b1:
                st.metric("💰 Số dư Bank (tại ref cuối)", f"{cutoff_balance:,.0f}")
            with col_b2:
                st.metric("📊 Số dư Account Sheet", f"{account_balance:,.0f}")
            with col_b3:
                if diff == 0:
                    st.metric("✅ Chênh lệch", "0 — KHỚP")
                else:
                    st.metric("⚠️ Chênh lệch", f"{diff:,.0f}")

            if diff == 0:
                st.success("✅ **Số dư KHỚP!** Sẵn sàng duyệt giao dịch mới.")
            else:
                st.error(f"❌ **Số dư LỆCH {diff:,.0f}** — Kiểm tra lại trước khi hạch toán!")
                if not st.checkbox("⚠️ Tôi đã kiểm tra, vẫn muốn tiếp tục duyệt", key="force_continue"):
                    return
        else:
            st.warning(f"⚠️ Không tìm thấy số dư cho `{raw_sheet_gsheet}` trong sheet Account")

    st.divider()

    # ═══════════════════════════════════════════════
    # HIỂN THỊ BẢNG GIAO DỊCH MỚI
    # ═══════════════════════════════════════════════
    transactions = new_transactions

    if not transactions:
        st.success("🎉 Tất cả giao dịch đã được hạch toán! Không còn giao dịch mới.")
        return

    st.subheader(f"🆕 {len(transactions)} giao dịch mới cần duyệt")

    dropdown_options = [BIG_ISSUE_OPTION] + project_sheets

    col_bulk1, col_bulk2 = st.columns([3, 1])
    with col_bulk1:
        bulk_sheet = st.selectbox(
            "⚡ Áp dụng nhanh 1 sheet cho tất cả dòng",
            ["-- Không áp dụng --"] + dropdown_options,
            key="bulk_sheet"
        )
    with col_bulk2:
        if st.button("Áp dụng", use_container_width=True, key="apply_bulk"):
            if bulk_sheet != "-- Không áp dụng --":
                for i in range(len(transactions)):
                    st.session_state[f"p2_sheet_{selected_key}_{i}"] = bulk_sheet
                st.rerun()

    st.divider()

    for i, tx in enumerate(transactions):
        color = "🟢" if tx['direction'] == 'THU' else "🔴"
        amount_fmt = f"{tx['amount']:,.0f}"
        sign = "+" if tx['direction'] == 'THU' else "-"

        c1, c2, c3, c4 = st.columns([1.2, 3.5, 2, 2.5])

        with c1:
            st.markdown(f"**{tx['date']}**")
            st.caption(f"{color} {tx['direction']}")

        with c2:
            desc_short = tx['desc'][:80] + ('...' if len(tx['desc']) > 80 else '')
            st.markdown(f"{desc_short}")
            if tx['counter_name']:
                st.caption(f"👤 {tx['counter_name']}")

        with c3:
            st.markdown(f"**{sign}{amount_fmt}**")

        with c4:
            default_key = f"p2_sheet_{selected_key}_{i}"
            default_idx = 0
            if default_key in st.session_state and st.session_state[default_key] in dropdown_options:
                default_idx = dropdown_options.index(st.session_state[default_key])

            st.selectbox(
                "Sheet",
                dropdown_options,
                index=default_idx,
                key=default_key,
                label_visibility="collapsed"
            )

        st.markdown("<hr style='margin:2px 0; border:none; border-top:1px solid #333'>", unsafe_allow_html=True)

    st.divider()

    col_s1, col_s2 = st.columns([1, 1])
    with col_s1:
        st.metric("Tổng giao dịch mới", len(transactions))
    with col_s2:
        total_thu = sum(tx['credit'] for tx in transactions if tx['direction'] == 'THU')
        total_chi = sum(tx['debit'] for tx in transactions if tx['direction'] == 'CHI')
        st.metric("THU / CHI", f"+{total_thu:,.0f} / -{total_chi:,.0f}")

    if st.button("✅ Duyệt & Hạch toán TẤT CẢ", type="primary", use_container_width=True):
        progress_bar = st.progress(0, text="Đang hạch toán...")
        success_count = 0
        error_list = []

        for i, tx in enumerate(transactions):
            try:
                sheet_key = f"p2_sheet_{selected_key}_{i}"
                selected_project = st.session_state.get(sheet_key, dropdown_options[0])

                raw_row = build_raw_row(tx, raw_sheet_gsheet, spreadsheet)
                append_to_raw_sheet(spreadsheet, raw_sheet_gsheet, raw_row)

                delta = tx['credit'] if tx['direction'] == 'THU' else -tx['debit']

                if selected_project == BIG_ISSUE_OPTION:
                    update_big_issue(spreadsheet, delta)
                else:
                    append_to_project_sheet(spreadsheet, selected_project,
                                           tx['date'], tx['desc'], -delta)

                cell_addr, _ = get_account_cell(spreadsheet, raw_sheet_gsheet)
                if cell_addr:
                    update_account_balance(spreadsheet, cell_addr, delta)

                success_count += 1

            except Exception as e:
                error_list.append(f"Dòng {i+1}: {str(e)}")

            progress_bar.progress((i + 1) / len(transactions),
                                  text=f"Đang hạch toán... {i+1}/{len(transactions)}")

            import time
            time.sleep(0.3)

        progress_bar.progress(1.0, text="Hoàn tất!")

        if success_count > 0:
            st.success(f"✅ Đã hạch toán thành công **{success_count}/{len(transactions)}** giao dịch vào **{raw_sheet_gsheet}**")
        if error_list:
            with st.expander(f"⚠️ {len(error_list)} lỗi", expanded=True):
                for e in error_list:
                    st.error(e)

# Thêm tab Phase 2 vào app
st.divider()
if st.toggle("📋 Mở Phase 2 — Duyệt lệnh & Hạch toán"):
    render_phase2()
